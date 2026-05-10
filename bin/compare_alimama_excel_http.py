#!/usr/bin/env python3
"""Compare Alimama HTTP/CDP-equivalent collection with an Excel reference."""

from __future__ import annotations

import argparse
import sys
from collections import Counter
from datetime import date, datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from tourism_automation.collectors.alimama_daily.collector import collect_alimama_daily
from tourism_automation.collectors.alimama_daily.excel_rules import BASE_FIELDS, SHEET_FIELDS, calculate_excel_metrics, calculate_wanxiangtai_2_total
from tourism_automation.collectors.alimama_daily.metrics import decimal_or_zero


CHANNEL_SHEETS = {
    "明星店铺": "star_store",
    "直通车": "tmall_express",
    "引力魔方": "gravity_rubiks_cube",
    "万相台": "wanxiangtai",
}

SOURCE_ORDER = ["超级短视频", "超级直播", "货品运营", "全站推广"]


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True)
    parser.add_argument("--start-date", required=True)
    parser.add_argument("--end-date", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    start = date.fromisoformat(args.start_date)
    end = date.fromisoformat(args.end_date)
    excel = _read_excel(Path(args.workbook))
    mismatches, missing, checked = _compare_range(excel, start=start, end=end)
    Path(args.output).write_text(_render_report(args.workbook, start, end, checked, missing, mismatches), encoding="utf-8")
    print(f"checked={checked} missing={len(missing)} mismatches={len(mismatches)} output={args.output}")
    return 0


def _read_excel(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    result: dict[str, Any] = {}
    for sheet, channel in CHANNEL_SHEETS.items():
        rows = {}
        previous_base = None
        for row in workbook[sheet].iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            biz_date = _date_key(row[0])
            base = dict(zip(BASE_FIELDS, row[1:9]))
            rows[biz_date] = calculate_excel_metrics(base, date_time=biz_date, sheet=channel, previous_row=previous_base)
            previous_base = base
        result[channel] = rows

    rows_by_date: dict[str, dict[str, dict[str, Any]]] = {}
    current_date = None
    for row in workbook["万相台2"].iter_rows(min_row=2, values_only=True):
        if row[0]:
            current_date = _date_key(row[0])
        if not current_date or not row[1]:
            continue
        source = row[1]
        if source == "小计":
            continue
        base = dict(zip(BASE_FIELDS, row[2:10]))
        metrics = calculate_excel_metrics(base, date_time=current_date, sheet="wanxiangtai_2")
        metrics["data_source"] = source
        rows_by_date.setdefault(current_date, {})[source] = metrics
    for biz_date, rows in rows_by_date.items():
        detail_rows = [rows.get(source, _zero_row(source, biz_date)) for source in SOURCE_ORDER]
        rows["小计"] = calculate_wanxiangtai_2_total(detail_rows, date_time=biz_date)
    result["wanxiangtai_2"] = rows_by_date
    return result


def _compare_range(excel: dict[str, Any], *, start: date, end: date):
    mismatches = []
    missing = []
    checked = 0
    current = start
    while current <= end:
        biz_date = current.isoformat()
        result = collect_alimama_daily(biz_date=biz_date)
        for sheet_name, channel in CHANNEL_SHEETS.items():
            excel_metrics = excel[channel].get(biz_date)
            http_metrics = result["channels"].get(channel)
            if not excel_metrics or not http_metrics:
                missing.append((biz_date, sheet_name, "missing row"))
                continue
            for field in SHEET_FIELDS[channel]:
                checked += 1
                if not _equalish(field, excel_metrics.get(field), http_metrics.get(field)):
                    mismatches.append((biz_date, sheet_name, field, excel_metrics.get(field), http_metrics.get(field)))
        for source in SOURCE_ORDER + ["小计"]:
            excel_metrics = excel["wanxiangtai_2"].get(biz_date, {}).get(source)
            http_metrics = result["wanxiangtai_2_rows"].get(source)
            if not excel_metrics or not http_metrics:
                missing.append((biz_date, f"万相台2/{source}", "missing row"))
                continue
            for field in SHEET_FIELDS["wanxiangtai_2"]:
                checked += 1
                if not _equalish(field, excel_metrics.get(field), http_metrics.get(field)):
                    mismatches.append((biz_date, f"万相台2/{source}", field, excel_metrics.get(field), http_metrics.get(field)))
        current += timedelta(days=1)
    return mismatches, missing, checked


def _render_report(workbook: str, start: date, end: date, checked: int, missing: list, mismatches: list) -> str:
    by_sheet = Counter(item[1] for item in mismatches)
    by_field = Counter(item[2] for item in mismatches)
    lines = [
        "# 阿里妈妈日报 HTTP/CDP/Excel 核对报告",
        "",
        f"- Excel: `{workbook}`",
        f"- 日期范围: `{start}` 至 `{end}`",
        f"- 检查字段数: `{checked}`",
        f"- 缺失项: `{len(missing)}`",
        f"- 不一致项: `{len(mismatches)}`",
        "",
        "## 口径说明",
        "",
        "- HTTP 使用本机 Chrome 登录态 Cookie。",
        "- 品销宝 CDP 页面实际请求包含 `rptCampaignList2.json`，当前代码仍会在报告中暴露与 Excel 的差异。",
        "- 万相台2 小计按当天 4 条明细重新汇总，不读取 Excel 小计缓存值。",
        "- 引力魔方 CPM 按 Excel 公式的上一行依赖计算参考值。",
        "",
        "## 差异汇总",
        "",
    ]
    if by_sheet:
        lines.extend([f"- {sheet}: {count}" for sheet, count in by_sheet.most_common()])
    else:
        lines.append("- 无差异")
    lines.extend(["", "## 高频字段", ""])
    if by_field:
        lines.extend([f"- {field}: {count}" for field, count in by_field.most_common(20)])
    else:
        lines.append("- 无差异")
    lines.extend(["", "## 差异明细（前 200 条）", "", "| 日期 | Sheet | 字段 | Excel | HTTP |", "|---|---|---|---:|---:|"])
    for biz_date, sheet, field, excel_value, http_value in mismatches[:200]:
        lines.append(f"| {biz_date} | {sheet} | {field} | {_fmt(excel_value)} | {_fmt(http_value)} |")
    if missing:
        lines.extend(["", "## 缺失项", ""])
        lines.extend([f"- `{item}`" for item in missing[:50]])
    return "\n".join(lines) + "\n"


def _equalish(field: str, excel_value: Any, http_value: Any) -> bool:
    if field == "data_source":
        return excel_value == http_value
    if field in {"cost", "sales", "cpc", "cpm", "asp", "cporder", "cpshopping_cart", "collection_cart_cost"}:
        return _quantize(excel_value, "0.01") == _quantize(http_value, "0.01")
    if field in {"ctr", "cvr", "cart_rate", "collection_cart_rate"}:
        return _quantize(excel_value, "0.000001") == _quantize(http_value, "0.000001")
    if field == "roi":
        return _quantize(excel_value, "0.0001") == _quantize(http_value, "0.0001")
    if field in {"imp", "click", "order_count", "shopping_cart", "bookmark_product", "bookmark_store", "collection_cart_count"}:
        return int(decimal_or_zero(excel_value)) == int(decimal_or_zero(http_value))
    return excel_value == http_value


def _quantize(value: Any, places: str) -> Decimal:
    return decimal_or_zero(value).quantize(Decimal(places), rounding=ROUND_HALF_UP)


def _date_key(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def _zero_row(source: str, biz_date: str) -> dict[str, Any]:
    metrics = calculate_excel_metrics(dict.fromkeys(BASE_FIELDS, 0), date_time=biz_date, sheet="wanxiangtai_2")
    metrics["data_source"] = source
    return metrics


def _fmt(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


if __name__ == "__main__":
    raise SystemExit(main())
