import importlib.util
import unittest
from decimal import Decimal
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parent.parent


def load_export_module():
    spec = importlib.util.spec_from_file_location(
        "export_daily_database_excel",
        ROOT / "skills/skills/openclaw-daily-database-excel/scripts/export_daily_database_excel.py",
    )
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module


class ExportDailyDatabaseExcelTests(unittest.TestCase):
    def test_convert_numeric_text_values(self):
        module = load_export_module()

        self.assertEqual((12, None), module._convert_numeric_text("12"))
        self.assertEqual((1234.5, None), module._convert_numeric_text("1,234.50"))
        self.assertEqual((1.98, None), module._convert_numeric_text("￥1.98"))
        self.assertEqual((0.0378, "0.00%"), module._convert_numeric_text("3.78%"))
        self.assertEqual((2.5, None), module._convert_numeric_text(Decimal("2.50")))
        self.assertEqual(("延迟统计", None), module._convert_numeric_text("延迟统计"))
        self.assertEqual(("2026-05-19", None), module._convert_numeric_text("2026-05-19"))

    def test_write_sheet_writes_numeric_strings_as_numbers(self):
        module = load_export_module()
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)

        module.write_sheet(
            workbook,
            "指标",
            ["点击率", "点击成本", "备注"],
            [("3.78%", "￥1.98", "延迟统计")],
            set(),
        )

        worksheet = workbook["指标"]
        self.assertEqual(0.0378, worksheet["A2"].value)
        self.assertEqual("0.00%", worksheet["A2"].number_format)
        self.assertEqual("n", worksheet["A2"].data_type)
        self.assertEqual(1.98, worksheet["B2"].value)
        self.assertEqual("n", worksheet["B2"].data_type)
        self.assertEqual("延迟统计", worksheet["C2"].value)
        self.assertEqual("s", worksheet["C2"].data_type)

    def test_apply_standard_table_style_formats_cells(self):
        module = load_export_module()
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.append(["字段", "值"])
        worksheet.append(["点击率", "3.78%"])

        module.apply_standard_table_style(workbook)

        for row in worksheet.iter_rows():
            for cell in row:
                self.assertEqual("center", cell.alignment.horizontal)
                self.assertEqual("center", cell.alignment.vertical)
                self.assertEqual("thin", cell.border.left.style)
        self.assertEqual("FF305496", worksheet["A1"].fill.fgColor.rgb)
        self.assertEqual("FFF2F2F2", worksheet["A1"].font.color.rgb)
        self.assertTrue(worksheet["A1"].font.bold)
        self.assertEqual(41.4, worksheet.row_dimensions[1].height)


if __name__ == "__main__":
    unittest.main()
