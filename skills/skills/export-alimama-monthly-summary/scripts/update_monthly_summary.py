#!/usr/bin/env python3
from __future__ import annotations

import re
import sys
from copy import copy
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.formula.translate import Translator


DEFAULT_WORKBOOK = Path("/mnt/c/Users/Gzk/Desktop/2026年阿里妈妈投放每日数据-更新到5.8(2).xlsx")
DEFAULT_OUTPUT_DIR = Path("/mnt/c/Users/Gzk/Desktop")
SUMMARY_SHEETS = ("月数据汇总-", "月数据汇总")
CHANNELS = ("明星店铺", "直通车", "引力魔方", "万相台")
METRIC_KEYS = {
    "cost": ("cost", "花费"),
    "imp": ("imp", "展示", "观看量"),
    "click": ("click", "点击"),
    "order": ("order", "订单"),
    "sales": ("sales", "销量", "gmv"),
    "shopping_cart": ("shoppingcart", "加入购物车"),
    "bookmark_product": ("bookmark-product", "宝贝收藏"),
    "bookmark_store": ("bookmark-store", "店铺收藏"),
}
SUMMARY_METRIC_HEADERS = {
    "cost": ("cost", "花费"),
    "imp": ("imp", "展示", "观看量"),
    "click": ("click", "点击"),
    "order": ("order", "订单"),
    "sales": ("sales", "销量", "gmv"),
    "shopping_cart": ("shoppingcart", "加入购物车"),
    "bookmark_product": ("bookmark-product", "宝贝收藏"),
    "bookmark_store": ("bookmark-store", "店铺收藏"),
    "bookmark_total": ("总收藏数",),
}
CHINESE_MONTHS = {
    "一": 1,
    "二": 2,
    "三": 3,
    "四": 4,
    "五": 5,
    "六": 6,
    "七": 7,
    "八": 8,
    "九": 9,
    "十": 10,
    "十一": 11,
    "十二": 12,
}


def normalize(value) -> str:
    return re.sub(r"\s+", "", str(value or "")).lower()


def parse_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            return date(1899, 12, 30) + timedelta(days=int(value))
        except OverflowError:
            return None
    if isinstance(value, str):
        text = value.strip()
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                pass
    return None


def parse_period(text: str) -> tuple[date, date]:
    start_text, end_text = re.split(r"[:~至]", text, maxsplit=1)
    return datetime.strptime(start_text.strip(), "%Y-%m-%d").date(), datetime.strptime(end_text.strip(), "%Y-%m-%d").date()


def period_for_report_month(year: int, month: int) -> tuple[date, date]:
    if month == 1:
        return date(year, 1, 1), date(year, 1, 20)
    if month == 12:
        return date(year, 11, 21), date(year, 12, 31)
    return date(year, month - 1, 21), date(year, month, 20)


def period_from_user_text(text: str) -> tuple[date, date]:
    text = text.strip()
    if re.search(r"\d{4}-\d{2}-\d{2}\s*[:~至]\s*\d{4}-\d{2}-\d{2}", text):
        return parse_period(text)

    year_match = re.search(r"(20\d{2})\s*年?", text)
    year = int(year_match.group(1)) if year_match else 2026

    numeric_month = re.search(r"(?<!\d)(1[0-2]|[1-9])\s*月", text)
    if numeric_month:
        return period_for_report_month(year, int(numeric_month.group(1)))

    for month_text in sorted(CHINESE_MONTHS, key=len, reverse=True):
        if f"{month_text}月" in text:
            return period_for_report_month(year, CHINESE_MONTHS[month_text])

    raise ValueError(f"无法从输入识别月份或周期：{text!r}")


def period_label(start: date, end: date) -> str:
    return f"{start.month}月{start.day}号-{end.month}月{end.day}号"


def period_from_label(label: str) -> tuple[date, date]:
    match = re.search(r"(\d+)月(\d+)号-(\d+)月(\d+)号", str(label))
    if not match:
        raise ValueError(f"无法从月汇总标题识别周期：{label!r}")
    start_month, start_day, end_month, end_day = map(int, match.groups())
    return date(2026, start_month, start_day), date(2026, end_month, end_day)


def next_period_after(label: str) -> tuple[date, date]:
    _, previous_end = period_from_label(label)
    start = previous_end + timedelta(days=1)
    if start == date(2026, 1, 21):
        end = date(2026, 2, 20)
    elif start == date(2026, 11, 21):
        end = date(2026, 12, 31)
    else:
        month = start.month + 1
        end_year = start.year
        if month == 13:
            month = 1
            end_year += 1
        end = date(end_year, month, 20)
    return start, end


def numeric(value) -> float:
    if value in (None, "", "-"):
        return 0
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip().replace(",", "").replace("￥", "").replace("¥", "")
    if text.endswith("%"):
        text = text[:-1]
    try:
        return float(text)
    except ValueError:
        return 0


def previous_period(start: date) -> tuple[date, date]:
    previous_end = start - timedelta(days=1)
    if previous_end == date(start.year, 12, 31):
        return date(start.year, 11, 21), previous_end
    if previous_end.day == 20:
        if previous_end.month == 1:
            return date(previous_end.year, 1, 1), previous_end
        return date(previous_end.year, previous_end.month - 1, 21), previous_end
    return period_for_report_month(previous_end.year, previous_end.month)


def find_summary_sheet(wb):
    for name in SUMMARY_SHEETS:
        if name in wb.sheetnames:
            return wb[name]
    raise ValueError("找不到月数据汇总 sheet")


def find_last_block(ws) -> int:
    candidates = []
    for row in range(1, ws.max_row + 1):
        value = ws.cell(row, 1).value
        if isinstance(value, str) and "号-" in value and "Cost" in str(ws.cell(row, 2).value):
            if all(
                ws.cell(row + offset, 1).value == channel
                for offset, channel in enumerate((*CHANNELS, "总计"), start=1)
            ):
                candidates.append(row)
    if not candidates:
        raise ValueError("找不到可复制的月度汇总块")
    return candidates[-1]


def find_template_block(wb, ws) -> int:
    value_wb = openpyxl.load_workbook(DEFAULT_WORKBOOK, data_only=True)
    value_ws = value_wb[ws.title]
    valid_blocks = []
    for row in range(1, ws.max_row + 1):
        label = ws.cell(row, 1).value
        if not (isinstance(label, str) and "号-" in label and "Cost" in str(ws.cell(row, 2).value)):
            continue
        if not all(
            ws.cell(row + offset, 1).value == channel
            for offset, channel in enumerate((*CHANNELS, "总计"), start=1)
        ):
            continue
        try:
            start, end = period_from_label(label)
        except ValueError:
            continue

        columns = summary_metric_columns(ws, row)
        if not set(SUMMARY_METRIC_HEADERS).issubset(columns):
            continue
        mismatches = 0
        for offset, channel in enumerate(CHANNELS, start=1):
            totals, _ = sum_channel(wb[channel], start, end)
            expected_values = {
                "cost": round(totals["cost"], 2),
                "imp": int(totals["imp"]),
                "click": int(totals["click"]),
                "order": int(totals["order"]),
                "sales": round(totals["sales"], 2),
                "shopping_cart": int(totals["shopping_cart"]),
                "bookmark_product": int(totals["bookmark_product"]),
                "bookmark_store": int(totals["bookmark_store"]),
                "bookmark_total": int(totals["bookmark_product"] + totals["bookmark_store"]),
            }
            for key, expected in expected_values.items():
                col = columns[key]
                if abs(numeric(value_ws.cell(row + offset, col).value) - expected) >= 0.01:
                    mismatches += 1
        if mismatches == 0:
            valid_blocks.append(row)

    return valid_blocks[-1] if valid_blocks else find_last_block(ws)


def metric_columns(ws) -> dict[str, int]:
    headers = [normalize(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
    result = {}
    for key, aliases in METRIC_KEYS.items():
        for idx, header in enumerate(headers, start=1):
            if any(alias.lower() in header for alias in aliases):
                result[key] = idx
                break
    return result


def summary_metric_columns(ws, header_row: int) -> dict[str, int]:
    headers = [normalize(ws.cell(header_row, col).value) for col in range(1, ws.max_column + 1)]
    result = {}
    for key, aliases in SUMMARY_METRIC_HEADERS.items():
        for idx, header in enumerate(headers, start=1):
            if any(alias.lower() in header for alias in aliases):
                result[key] = idx
                break
    return result


def sum_channel(ws, start: date, end: date) -> tuple[dict[str, float], int]:
    cols = metric_columns(ws)
    totals = {key: 0 for key in METRIC_KEYS}
    count = 0
    for row in range(2, ws.max_row + 1):
        row_date = parse_date(ws.cell(row, 1).value)
        if row_date is None or not (start <= row_date <= end):
            continue
        count += 1
        for key, col in cols.items():
            totals[key] += numeric(ws.cell(row, col).value)
    return totals, count


def copy_row(ws, src_row: int, dst_row: int) -> None:
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    for col in range(1, ws.max_column + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        dst._style = copy(src._style)
        dst.number_format = src.number_format
        dst.alignment = copy(src.alignment)
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        if isinstance(src.value, str) and src.value.startswith("="):
            dst.value = Translator(src.value, origin=src.coordinate).translate_formula(dst.coordinate)
        else:
            dst.value = src.value


def copy_row_style(src_ws, dst_ws, src_row: int, dst_row: int, header_row: bool = False) -> None:
    dst_ws.row_dimensions[dst_row].height = src_ws.row_dimensions[src_row].height
    for col in range(1, src_ws.max_column + 1):
        src = src_ws.cell(src_row, col)
        dst = dst_ws.cell(dst_row, col)
        dst._style = copy(src._style)
        dst.number_format = src.number_format
        dst.alignment = copy(src.alignment)
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        if header_row or col == 1:
            dst.value = src.value


def fill_base_values(wb, ws, period: tuple[date, date], block_start: int, columns: dict[str, int]) -> dict[str, int]:
    start, end = period
    counts = {}
    for offset, channel in enumerate(CHANNELS, start=1):
        totals, count = sum_channel(wb[channel], start, end)
        counts[channel] = count
        row = block_start + offset
        ws.cell(row, 1).value = channel
        values = {
            "cost": round(totals["cost"], 2),
            "imp": int(totals["imp"]),
            "click": int(totals["click"]),
            "order": int(totals["order"]),
            "sales": round(totals["sales"], 2),
            "shopping_cart": int(totals["shopping_cart"]),
            "bookmark_product": int(totals["bookmark_product"]),
            "bookmark_store": int(totals["bookmark_store"]),
            "bookmark_total": int(totals["bookmark_product"] + totals["bookmark_store"]),
        }
        for key, value in values.items():
            if key in columns:
                ws.cell(row, columns[key]).value = value
    return counts


def apply_summary_formulas(ws, block_start: int, previous_block_start: int | None = None) -> None:
    total_row = block_start + 5
    ws.cell(total_row, 1).value = "总计"
    for row in range(block_start + 1, block_start + 6):
        if row == total_row:
            for col in range(2, 11):
                letter = ws.cell(row, col).column_letter
                ws.cell(row, col).value = f"=SUM({letter}{block_start + 1}:{letter}{block_start + 4})"
        else:
            ws.cell(row, 10).value = f"=H{row}+I{row}"

        ws.cell(row, 11).value = f"=D{row}/C{row}"
        ws.cell(row, 12).value = f"=B{row}/D{row}"
        ws.cell(row, 13).value = f"=(B{row}/C{row})*1000"
        ws.cell(row, 14).value = f"=F{row}/B{row}"
        ws.cell(row, 15).value = f"=E{row}/D{row}"
        ws.cell(row, 16).value = f"=IF(E{row}=0,0,F{row}/E{row})"
        ws.cell(row, 17).value = f"=IF(E{row}=0,0,B{row}/E{row})"
        ws.cell(row, 18).value = f"=B{row}/G{row}"
        ws.cell(row, 19).value = f"=G{row}+J{row}"
        ws.cell(row, 21).value = f"=B{row}/S{row}"
        ws.cell(row, 28).value = f"=B{row}/B{total_row}"
        if row == total_row:
            ws.cell(row, 29).value = f"=B{row}/F{row}"

        if previous_block_start is not None:
            prev = previous_block_start + (row - block_start)
            ws.cell(row, 20).value = f'=IF(S{prev}=0,"",(S{row}-S{prev})/S{prev})'
            ws.cell(row, 22).value = f'=IF(B{prev}=0,"",(B{row}-B{prev})/B{prev})'
            ws.cell(row, 23).value = f'=IF($D${total_row}=0,"",D{row}/$D${total_row})'
            ws.cell(row, 24).value = f'=IF(G{prev}=0,"",(G{row}-G{prev})/G{prev})'
            ws.cell(row, 25).value = f'=IF(F{prev}=0,"",(F{row}-F{prev})/F{prev})'
            ws.cell(row, 26).value = f'=IF(R{prev}=0,"",(R{row}-R{prev})/R{prev})'
            ws.cell(row, 27).value = f'=IF(U{prev}=0,"",(U{row}-U{prev})/U{prev})'


def export_new_workbook(wb, period: tuple[date, date], output_path: Path | None = None) -> Path:
    source_ws = find_summary_sheet(wb)
    template_start = find_template_block(wb, source_ws)
    start, end = period
    prev_period = previous_period(start)
    output_path = output_path or DEFAULT_OUTPUT_DIR / f"alimama_{end.month}月月数据汇总.xlsx"

    new_wb = Workbook()
    ws = new_wb.active
    ws.title = f"{end.month}月月数据汇总"

    for col in range(1, source_ws.max_column + 1):
        letter = source_ws.cell(1, col).column_letter
        ws.column_dimensions[letter].width = source_ws.column_dimensions[letter].width

    for block_start, label_period in ((1, period), (8, prev_period)):
        for offset in range(6):
            copy_row_style(source_ws, ws, template_start + offset, block_start + offset, header_row=(offset == 0))
            if offset != 0:
                ws.cell(block_start + offset, 1).value = source_ws.cell(template_start + offset, 1).value
        ws.cell(block_start, 1).value = period_label(*label_period)

    columns = summary_metric_columns(ws, 1)
    counts = fill_base_values(wb, ws, period, 1, columns)
    fill_base_values(wb, ws, prev_period, 8, columns)
    apply_summary_formulas(ws, 8)
    apply_summary_formulas(ws, 1, previous_block_start=8)

    for row in range(8, 14):
        ws.row_dimensions[row].hidden = True
    ws.freeze_panes = "B2"
    new_wb.save(output_path)

    print(f"output={output_path}")
    print(f"source_workbook={DEFAULT_WORKBOOK}")
    print(f"template_row={template_start}")
    print(f"period={start.isoformat()}:{end.isoformat()}")
    print(f"previous_period={prev_period[0].isoformat()}:{prev_period[1].isoformat()}")
    for channel, count in counts.items():
        print(f"{channel}\t{count} rows")
    return output_path


def append_block(wb, period: tuple[date, date]) -> None:
    ws = find_summary_sheet(wb)
    template_start = find_template_block(wb, ws)
    start, end = period
    target_start = ws.max_row + 2

    for offset in range(6):
        copy_row(ws, template_start + offset, target_start + offset)

    ws.cell(target_start, 1).value = period_label(start, end)
    columns = summary_metric_columns(ws, target_start)
    counts = {}
    for offset, channel in enumerate(CHANNELS, start=1):
        totals, count = sum_channel(wb[channel], start, end)
        counts[channel] = count
        row = target_start + offset
        ws.cell(row, 1).value = channel
        values = {
            "cost": round(totals["cost"], 2),
            "imp": int(totals["imp"]),
            "click": int(totals["click"]),
            "order": int(totals["order"]),
            "sales": round(totals["sales"], 2),
            "shopping_cart": int(totals["shopping_cart"]),
            "bookmark_product": int(totals["bookmark_product"]),
            "bookmark_store": int(totals["bookmark_store"]),
            "bookmark_total": int(totals["bookmark_product"] + totals["bookmark_store"]),
        }
        for key, value in values.items():
            if key in columns:
                ws.cell(row, columns[key]).value = value

    total_row = target_start + 5
    ws.cell(total_row, 1).value = "总计"
    for col in columns.values():
        letter = ws.cell(total_row, col).column_letter
        ws.cell(total_row, col).value = f"=SUM({letter}{target_start + 1}:{letter}{target_start + 4})"

    print(f"workbook={DEFAULT_WORKBOOK}")
    print(f"summary_sheet={ws.title}")
    print(f"template_row={template_start}")
    print(f"period={start.isoformat()}:{end.isoformat()}")
    print(f"new_block_row={target_start}")
    for channel, count in counts.items():
        print(f"{channel}\t{count} rows")


def main() -> int:
    workbook_path = DEFAULT_WORKBOOK
    period_arg = " ".join(sys.argv[1:]).strip() if len(sys.argv) > 1 else None
    wb = openpyxl.load_workbook(workbook_path)
    ws = find_summary_sheet(wb)
    period = period_from_user_text(period_arg) if period_arg else next_period_after(ws.cell(find_template_block(wb, ws), 1).value)
    export_new_workbook(wb, period)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
