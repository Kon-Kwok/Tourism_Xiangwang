#!/usr/bin/env python3
from __future__ import annotations

import argparse
import os
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

import openpyxl
import pymysql


PROJECT_ROOT = Path(__file__).resolve().parents[5]
DEFAULT_TABLES = (
    ("customer_service_data_daily", "日期", "赤兔-人均日接入"),
    ("customer_service_performance_summary", "date_time", "赤兔-每周店铺个人数据"),
    ("customer_service_performance_workload_analysis", "date_time", "赤兔-客服数据23年新"),
    ("shop_daily_key_data", "日期", "店铺日度关键数据"),
    ("shop_data_daily_registration", "日期", "店铺每日登记"),
    ("star_store", "date_time", "阿里妈妈-明星店铺"),
    ("tmall_express", "date_time", "阿里妈妈-直通车"),
    ("gravity_rubiks_cube", "date_time", "阿里妈妈-引力魔方"),
    ("wanxiangtai", "date_time", "阿里妈妈-万相台"),
    ("wanxiangtai_2", "date_time", "阿里妈妈-万相台2"),
)
TABLE_DISPLAY_NAMES = {table_name: display_name for table_name, _, display_name in DEFAULT_TABLES}
DATE_COLUMN_CANDIDATES = ("日期", "date_time", "order_date", "biz_date", "collection_date")
EXCLUDE_COLUMNS = {"created_at", "updated_at"}


def load_env() -> None:
    env_path = PROJECT_ROOT / ".env"
    if not env_path.exists():
        return

    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def parse_date(value: str | None) -> str:
    if value is None:
        return date.today().isoformat()
    return datetime.strptime(value, "%Y-%m-%d").date().isoformat()


def connect(args):
    return pymysql.connect(
        host=args.host,
        port=args.port,
        user=args.user,
        password=args.password,
        database=args.database,
        charset="utf8mb4",
        connect_timeout=5,
    )


def existing_default_tables(cursor, database: str) -> list[tuple[str, str, str]]:
    available = []
    for table_name, date_column, display_name in DEFAULT_TABLES:
        cursor.execute(
            """
            SELECT COUNT(*)
            FROM information_schema.COLUMNS
            WHERE TABLE_SCHEMA = %s
              AND TABLE_NAME = %s
              AND COLUMN_NAME = %s
            """,
            (database, table_name, date_column),
        )
        if cursor.fetchone()[0]:
            available.append((table_name, date_column, display_name))
    return available


def all_date_tables(cursor, database: str) -> list[tuple[str, str, str]]:
    placeholders = ", ".join(["%s"] * len(DATE_COLUMN_CANDIDATES))
    cursor.execute(
        f"""
        SELECT TABLE_NAME, COLUMN_NAME
        FROM information_schema.COLUMNS
        WHERE TABLE_SCHEMA = %s
          AND COLUMN_NAME IN ({placeholders})
        ORDER BY TABLE_NAME,
          FIELD(COLUMN_NAME, {placeholders})
        """,
        (database, *DATE_COLUMN_CANDIDATES, *DATE_COLUMN_CANDIDATES),
    )
    seen = set()
    tables = []
    for table_name, column_name in cursor.fetchall():
        if table_name in seen:
            continue
        seen.add(table_name)
        tables.append((table_name, column_name, TABLE_DISPLAY_NAMES.get(table_name, table_name)))
    return tables


def _clean_columns_rows(columns, rows, date_column):
    keep_idx = [i for i, col in enumerate(columns) if col not in EXCLUDE_COLUMNS]
    columns = [columns[i] for i in keep_idx]
    rows = [tuple(row[i] for i in keep_idx) for row in rows]
    new_order = []
    for head in ("id", date_column):
        try:
            idx = columns.index(head)
            new_order.append(idx)
        except ValueError:
            pass
    for i in range(len(columns)):
        if i not in new_order:
            new_order.append(i)
    columns = [columns[i] for i in new_order]
    rows = [tuple(row[i] for i in new_order) for row in rows]
    return columns, rows


def fetch_table(cursor, database: str, table_name: str, date_column: str, biz_date: str):
    cursor.execute(f"SELECT * FROM `{database}`.`{table_name}` WHERE `{date_column}` = %s", (biz_date,))
    columns = [column[0] for column in cursor.description]
    rows = cursor.fetchall()
    return _clean_columns_rows(columns, rows, date_column)


def fetch_table_range(cursor, database: str, table_name: str, date_column: str, start_date: str, end_date: str):
    cursor.execute(
        f"SELECT * FROM `{database}`.`{table_name}` WHERE `{date_column}` BETWEEN %s AND %s ORDER BY `{date_column}`, `id`",
        (start_date, end_date),
    )
    columns = [column[0] for column in cursor.description]
    rows = cursor.fetchall()
    return _clean_columns_rows(columns, rows, date_column)


def safe_sheet_name(name: str, used: set[str]) -> str:
    base = name[:31]
    candidate = base
    index = 2
    while candidate in used:
        suffix = f"_{index}"
        candidate = f"{base[:31 - len(suffix)]}{suffix}"
        index += 1
    used.add(candidate)
    return candidate


def write_sheet(workbook, sheet_name: str, columns: Iterable[str], rows: Iterable[tuple], used: set[str]) -> int:
    worksheet = workbook.create_sheet(safe_sheet_name(sheet_name, used))
    worksheet.append(list(columns))
    count = 0
    for row in rows:
        worksheet.append(list(row))
        count += 1
    worksheet.freeze_panes = "A2"
    return count


COLUMN_FORMAT_RULES: dict[str, dict[str, str]] = {
    "赤兔-人均日接入": {
        "回复率": "0.00%",
        "询单最终付款成功率": "0.00%",
        "评价发送率": "0.00%",
        "客户满意比": "0.0000",
        "很满意": "0",
        "满意": "0",
        "一般": "0",
        "不满意": "0",
        "很不满意": "0",
    },
    "赤兔-每周店铺个人数据": {
        "询单人数": "0",
    },
    "赤兔-客服数据23年新": {
        "未回复人数": "0.00%",
        "旺旺回复率": "0.00%",
    },
    "店铺每日登记": {
        "咨询转化率": "0.00%",
        "下单转化率": "0.00%",
    },
}


def _apply_cell_formats(workbook) -> None:
    for worksheet in workbook.worksheets:
        name = worksheet.title
        rules = COLUMN_FORMAT_RULES.get(name)
        if rules is None:
            continue

        header_row = worksheet[1]
        col_indices: dict[str, int] = {}
        for idx, cell in enumerate(header_row):
            if cell.value in rules:
                col_indices[cell.value] = idx

        if not col_indices:
            continue

        for row in worksheet.iter_rows(min_row=2):
            for col_name, col_idx in col_indices.items():
                cell = row[col_idx]
                fmt_code = rules[col_name]
                value = cell.value

                if value is None:
                    continue

                if isinstance(value, str):
                    try:
                        value = float(value)
                    except (ValueError, TypeError):
                        continue

                cell.value = value
                cell.number_format = fmt_code


def _handle_delay_chat_volume(workbook, summary: list) -> None:
    """替换最新日期的 chat_volume=0 为 '延迟统计'。"""
    ws = None
    for worksheet in workbook.worksheets:
        if worksheet.title == "店铺日度关键数据":
            ws = worksheet
            break
    if ws is None:
        return

    header_row = ws[1]
    date_col_idx, chat_col_idx = None, None
    for idx, cell in enumerate(header_row):
        if cell.value == "日期":
            date_col_idx = idx
        elif cell.value == "chat_volume":
            chat_col_idx = idx
    if date_col_idx is None or chat_col_idx is None:
        return

    dates = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = row[date_col_idx]
        if isinstance(d, date):
            dates.add(d)
    max_date = max(dates) if dates else None

    if max_date is None:
        return

    for row in ws.iter_rows(min_row=2):
        date_cell = row[date_col_idx]
        chat_cell = row[chat_col_idx]
        if isinstance(date_cell.value, date) and date_cell.value == max_date and chat_cell.value == 0:
            chat_cell.value = "延迟统计"
            chat_cell.number_format = "@"


def autosize_workbook(workbook) -> None:
    for worksheet in workbook.worksheets:
        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells[:100]:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 40)


def build_workbook(conn, args, biz_date: str, start_date: str = None, end_date: str = None):
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    used_sheet_names: set[str] = set()
    summary = []
    range_mode = start_date is not None and end_date is not None

    with conn.cursor() as cursor:
        tables = all_date_tables(cursor, args.database) if args.all_date_tables else existing_default_tables(cursor, args.database)

        for table_name, date_column, display_name in tables:
            if range_mode:
                columns, rows = fetch_table_range(cursor, args.database, table_name, date_column, start_date, end_date)
            else:
                columns, rows = fetch_table(cursor, args.database, table_name, date_column, biz_date)
            if not rows and not args.include_empty_tables:
                continue
            row_count = write_sheet(workbook, display_name, columns, rows, used_sheet_names)
            summary.append((display_name, date_column, row_count))

    overview = workbook.create_sheet("汇总", 0)
    if range_mode:
        overview.append(["日期范围", f"{start_date} 至 {end_date}", "", ""])
    else:
        overview.append(["日期", biz_date])
    overview.append([])
    overview.append(["表名", "日期字段", "行数"])
    for row in summary:
        overview.append(list(row))
    overview.freeze_panes = "A4"
    _apply_cell_formats(workbook)
    _handle_delay_chat_volume(workbook, summary)
    autosize_workbook(workbook)
    return workbook, summary


def main() -> int:
    load_env()
    parser = argparse.ArgumentParser(description="Export Xiangwang daily database rows to Excel")
    parser.add_argument("--date", help="Business date, format YYYY-MM-DD. Defaults to today.")
    parser.add_argument("--start", help="Start date for range export, format YYYY-MM-DD.")
    parser.add_argument("--end", help="End date for range export, format YYYY-MM-DD.")
    parser.add_argument("--output", help="Output xlsx path. Defaults to exports/daily_database_YYYY-MM-DD.xlsx")
    parser.add_argument("--host", default=os.environ.get("HOST", "127.0.0.1"))
    parser.add_argument("--port", type=int, default=int(os.environ.get("PORT", "3306")))
    parser.add_argument("--user", default=os.environ.get("USER", "remote_user"))
    parser.add_argument("--password", default=os.environ.get("PASS", "Tourism2024"))
    parser.add_argument("--database", default=os.environ.get("DATABASE", "Xiangwang"))
    parser.add_argument("--all-date-tables", action="store_true", help="Export every table with a recognized date column.")
    parser.add_argument("--include-empty-tables", action="store_true", help="Also create sheets for tables with no rows on the selected date.")
    args = parser.parse_args()

    range_mode = args.start is not None and args.end is not None
    if (args.start is not None) != (args.end is not None):
        parser.error("--start and --end must be used together")

    if range_mode:
        start_date = parse_date(args.start)
        end_date = parse_date(args.end)
        biz_date = None
    else:
        biz_date = parse_date(args.date)
        start_date = end_date = None

    if args.output:
        output_path = Path(args.output)
    elif range_mode:
        output_path = PROJECT_ROOT / "exports" / f"daily_database_{start_date}_{end_date}.xlsx"
    else:
        output_path = PROJECT_ROOT / "exports" / f"daily_database_{biz_date}.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        conn = connect(args)
    except pymysql.err.OperationalError:
        if "--host" in os.sys.argv or args.host in ("127.0.0.1", "localhost"):
            raise
        args.host = "127.0.0.1"
        conn = connect(args)
    try:
        workbook, summary = build_workbook(conn, args, biz_date=biz_date, start_date=start_date, end_date=end_date)
        workbook.save(output_path)
    finally:
        conn.close()

    print(f"output={output_path}")
    for table_name, date_column, row_count in summary:
        print(f"{table_name}\t{date_column}\t{row_count}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
